"""Build JSON data files from Platinum Redux v3.3 docs.xlsx + encounters.docx.

Run: python3 scripts/build_data.py

Source files are expected at:
  ../platinum redux/Platinum Redux v3.3 docs.xlsx
  ../platinum redux/Platinum Redux encounters v3.3.docx

Outputs JSON to data/.
"""

import json
import re
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
SRC = ROOT.parent / "platinum redux"
XLSX = SRC / "Platinum Redux v3.3 docs.xlsx"
DOCX = SRC / "Platinum Redux encounters v3.3.docx"
OUT = ROOT / "data"
OUT.mkdir(exist_ok=True)


def num(v):
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return v
    s = str(v).strip()
    if s == "" or s == "-":
        return None
    try:
        f = float(s)
        return int(f) if f.is_integer() else f
    except ValueError:
        return s


def slug(name):
    if not name:
        return ""
    s = str(name).strip().lower()
    s = re.sub(r"[^a-z0-9]+", "-", s)
    return s.strip("-")


def cell(ws, r, c):
    return ws.cell(r, c).value


def load_workbook():
    return openpyxl.load_workbook(XLSX, data_only=True)


def parse_pokemon(wb):
    ws = wb["Pokémon"]
    out = []
    for r in range(3, ws.max_row + 1):
        if cell(ws, r, 1) is None:
            continue
        out.append(
            {
                "id": int(cell(ws, r, 1)),
                "name": cell(ws, r, 2),
                "type1": cell(ws, r, 3),
                "type2": cell(ws, r, 4) if cell(ws, r, 4) != cell(ws, r, 3) else None,
                "stats": {
                    "hp": num(cell(ws, r, 5)),
                    "atk": num(cell(ws, r, 6)),
                    "def": num(cell(ws, r, 7)),
                    "spa": num(cell(ws, r, 8)),
                    "spd": num(cell(ws, r, 9)),
                    "spe": num(cell(ws, r, 10)),
                    "bst": num(cell(ws, r, 11)),
                },
                "ability1": cell(ws, r, 12),
                "ability2": cell(ws, r, 13),
                "evolve": cell(ws, r, 14),
                "wild": cell(ws, r, 15),
                "special": cell(ws, r, 16),
            }
        )
    return out


def parse_forms(wb):
    ws = wb["Forms"]
    out = []
    for r in range(3, ws.max_row + 1):
        if cell(ws, r, 1) is None and cell(ws, r, 2) is None:
            continue
        out.append(
            {
                "id": num(cell(ws, r, 1)),
                "name": cell(ws, r, 2),
                "type1": cell(ws, r, 3),
                "type2": cell(ws, r, 4) if cell(ws, r, 4) != cell(ws, r, 3) else None,
                "stats": {
                    "hp": num(cell(ws, r, 5)),
                    "atk": num(cell(ws, r, 6)),
                    "def": num(cell(ws, r, 7)),
                    "spa": num(cell(ws, r, 8)),
                    "spd": num(cell(ws, r, 9)),
                    "spe": num(cell(ws, r, 10)),
                    "bst": num(cell(ws, r, 11)),
                },
                "front": cell(ws, r, 12),
                "back": cell(ws, r, 13),
                "shiny": cell(ws, r, 14),
                "flavor": cell(ws, r, 15),
            }
        )
    return out


def parse_levelup(wb):
    """LVL Up sheet: alternating MOVE/LVL columns starting col 3."""
    ws = wb["LVL Up"]
    out = {}
    for r in range(3, ws.max_row + 1):
        pid = num(cell(ws, r, 1))
        name = cell(ws, r, 2)
        if pid is None and name is None:
            continue
        moves = []
        c = 3
        while c <= ws.max_column:
            mv = cell(ws, r, c)
            lvl = cell(ws, r, c + 1)
            if mv is not None and str(mv).strip():
                moves.append({"move": str(mv).strip(), "level": num(lvl)})
            c += 2
        if pid is not None:
            out[int(pid)] = moves
    return out


def parse_tm_learn(wb):
    """TM Learn structure:
      R2: single-copy location
      R3: infinite location
      R4: 'TM 1', 'TM 2', ...
      R5: TM move name
      R6: TM type
      R7+: Pokemon row -- col1 = id, col2 = name, col3+ = bool (learnable)
    """
    ws = wb["TM Learn"]

    tms = []  # list of {tm: 'TM 1', move: 'Body Slam', type: 'Normal', col, location, infinite}
    for c in range(3, ws.max_column + 1):
        tm_label = cell(ws, 4, c)
        move = cell(ws, 5, c)
        type_ = cell(ws, 6, c)
        if not move:
            continue
        tms.append(
            {
                "col": c,
                "tm": str(tm_label).strip() if tm_label else None,
                "move": str(move).strip(),
                "type": str(type_).strip() if type_ else None,
                "location": cell(ws, 2, c),
                "infinite": cell(ws, 3, c),
            }
        )

    learnsets = {}
    for r in range(7, ws.max_row + 1):
        pid = num(cell(ws, r, 1))
        if pid is None:
            continue
        learns = []
        for tm in tms:
            v = cell(ws, r, tm["col"])
            if v is True or (isinstance(v, str) and v.strip().lower() in ("y", "yes", "x", "true", "✓", "o", "1")):
                learns.append(tm["move"])
            elif isinstance(v, (int, float)) and v == 1:
                learns.append(tm["move"])
        learnsets[int(pid)] = learns

    return {
        "tms": [
            {
                "tm": t["tm"],
                "move": t["move"],
                "type": t["type"],
                "location": str(t["location"]) if t["location"] is not None else None,
                "infinite": str(t["infinite"]) if t["infinite"] is not None else None,
            }
            for t in tms
        ],
        "learnsets": learnsets,
    }


def parse_moves(wb):
    ws = wb["All Moves"]
    out = []
    for r in range(3, ws.max_row + 1):
        num_ = num(cell(ws, r, 1))
        name = cell(ws, r, 2)
        if name is None:
            continue
        out.append(
            {
                "id": num_,
                "name": name,
                "effect": cell(ws, r, 3),
                "category": cell(ws, r, 4),
                "type": cell(ws, r, 6),
                "power": num(cell(ws, r, 7)),
                "accuracy": num(cell(ws, r, 8)),
                "pp": num(cell(ws, r, 9)),
                "effect_pct": num(cell(ws, r, 10)),
                "priority": num(cell(ws, r, 11)),
                "tm": cell(ws, r, 12),
                "properties": cell(ws, r, 13),
                "avg_dmg": num(cell(ws, r, 14)),
            }
        )
    return out


def parse_type_chart(wb):
    ws = wb["Type Chart"]
    # Defending types in row 2 cols 3..
    defending = []
    for c in range(3, ws.max_column + 1):
        v = cell(ws, 2, c)
        if v:
            defending.append(str(v).strip())
    # Attacking types in col 2 starting row 3
    chart = {}
    for r in range(3, ws.max_row + 1):
        atk = cell(ws, r, 2)
        if not atk:
            continue
        atk = str(atk).strip()
        row_data = {}
        for i, dt in enumerate(defending):
            v = num(cell(ws, r, 3 + i))
            if v is None:
                v = 1
            row_data[dt] = v
        chart[atk] = row_data
    return {"types": defending, "chart": chart}


def parse_tc_changes(wb):
    ws = wb["TC Change"]
    out = []
    for r in range(3, ws.max_row + 1):
        atk = cell(ws, r, 1)
        defd = cell(ws, r, 2)
        if not atk or not defd:
            continue
        out.append(
            {
                "attacking": str(atk).strip(),
                "defending": str(defd).strip(),
                "modifier": num(cell(ws, r, 3)),
                "explanation": cell(ws, r, 4),
            }
        )
    return out


def parse_trainers(wb):
    """Each trainer can have multiple Pokemon (one per row). Group consecutive
    rows with same Area+Name."""
    ws = wb["Trainers"]
    rows = []
    last_area = None
    last_name = None
    for r in range(3, ws.max_row + 1):
        area = cell(ws, r, 1)
        name = cell(ws, r, 2)
        pkmn = cell(ws, r, 3)
        if pkmn is None:
            continue
        # Inherit area/name if blank (continuation of previous trainer)
        if area is None and name is None:
            area = last_area
            name = last_name
        else:
            last_area = area or last_area
            last_name = name or last_name
        rows.append(
            {
                "area": area,
                "name": name,
                "pokemon": pkmn,
                "type1": cell(ws, r, 4),
                "type2": cell(ws, r, 5) if cell(ws, r, 5) != cell(ws, r, 4) else None,
                "level": num(cell(ws, r, 6)),
                "ability": cell(ws, r, 7),
                "item": cell(ws, r, 8),
                "moves": [
                    cell(ws, r, 9),
                    cell(ws, r, 10),
                    cell(ws, r, 11),
                    cell(ws, r, 12),
                ],
                "dspre": num(cell(ws, r, 13)),
            }
        )

    # Group into trainers by (area, name) preserving order
    trainers = []
    current = None
    for row in rows:
        key = (row["area"], row["name"])
        if current is None or (current["area"], current["name"]) != key:
            current = {
                "area": row["area"],
                "name": row["name"],
                "team": [],
            }
            trainers.append(current)
        current["team"].append(
            {
                "pokemon": row["pokemon"],
                "type1": row["type1"],
                "type2": row["type2"],
                "level": row["level"],
                "ability": row["ability"],
                "item": row["item"],
                "moves": [m for m in row["moves"] if m and str(m).strip() not in ("-", "")],
                "dspre": row["dspre"],
            }
        )
    return trainers


def parse_bosses(wb):
    ws = wb["Bosses"]
    rows = []
    last_id = None
    last_boss = None
    for r in range(3, ws.max_row + 1):
        id_ = num(cell(ws, r, 1))
        boss = cell(ws, r, 2)
        pkmn = cell(ws, r, 3)
        if pkmn is None:
            continue
        if id_ is None and boss is None:
            id_ = last_id
            boss = last_boss
        else:
            last_id = id_ if id_ is not None else last_id
            last_boss = boss if boss is not None else last_boss
        rows.append(
            {
                "id": id_,
                "boss": boss,
                "pokemon": pkmn,
                "type1": cell(ws, r, 4),
                "type2": cell(ws, r, 5) if cell(ws, r, 5) != cell(ws, r, 4) else None,
                "level": num(cell(ws, r, 6)),
                "nature": cell(ws, r, 7),
                "ivs": cell(ws, r, 8),
                "speed": num(cell(ws, r, 9)),
                "nature_hc": cell(ws, r, 10),
                "ivs_hc": cell(ws, r, 11),
                "speed_hc": num(cell(ws, r, 12)),
                "ability": cell(ws, r, 13),
                "item": cell(ws, r, 14),
                "moves": [
                    cell(ws, r, 15),
                    cell(ws, r, 16),
                    cell(ws, r, 17),
                    cell(ws, r, 18),
                ],
                "dspre": num(cell(ws, r, 19)),
            }
        )

    bosses = []
    current = None
    for row in rows:
        key = (row["id"], row["boss"])
        if current is None or (current["id"], current["name"]) != key:
            current = {"id": row["id"], "name": row["boss"], "team": []}
            bosses.append(current)
        current["team"].append(
            {
                "pokemon": row["pokemon"],
                "type1": row["type1"],
                "type2": row["type2"],
                "level": row["level"],
                "nature": row["nature"],
                "ivs": row["ivs"],
                "speed": row["speed"],
                "nature_hc": row["nature_hc"],
                "ivs_hc": row["ivs_hc"],
                "speed_hc": row["speed_hc"],
                "ability": row["ability"],
                "item": row["item"],
                "moves": [m for m in row["moves"] if m and str(m).strip() not in ("-", "")],
                "dspre": row["dspre"],
            }
        )
    return bosses


def parse_encounters(wb):
    """Encounters v3.3 sheet: row 1 has area names spanning 2 columns each.
    Row 2 has Pokemon/% headers. Rows 3+ list Pokemon and probability."""
    ws = wb["Encounters v3.3"]
    # Areas: row 1, every 2 columns starting col 2
    areas = []
    c = 2
    while c <= ws.max_column:
        v = cell(ws, 1, c)
        if v:
            areas.append({"col": c, "name": str(v).strip()})
        c += 2

    out = []
    for area in areas:
        c = area["col"]
        encs = []
        for r in range(3, ws.max_row + 1):
            pkmn = cell(ws, r, c)
            pct = cell(ws, r, c + 1)
            if pkmn is None:
                continue
            encs.append({"pokemon": str(pkmn).strip(), "pct": num(pct)})
        if encs:
            out.append({"area": area["name"], "encounters": encs})
    return out


def parse_battle_items(wb):
    ws = wb["Battle Items"]
    out = []
    for r in range(3, ws.max_row + 1):
        item = cell(ws, r, 1)
        if not item:
            continue
        out.append(
            {
                "item": item,
                "effect": cell(ws, r, 2),
                "value": num(cell(ws, r, 3)),
                "location": cell(ws, r, 4),
            }
        )
    return out


def parse_walkthrough(wb):
    ws = wb["Walkthrough"]
    out = []
    last_area = None
    for r in range(3, ws.max_row + 1):
        area = cell(ws, r, 1)
        item = cell(ws, r, 2)
        rest = cell(ws, r, 3)
        if area:
            last_area = area
        if not item:
            continue
        out.append(
            {
                "area": last_area,
                "item": item,
                "restriction": rest,
            }
        )
    return out


def parse_tm_data(wb):
    ws = wb["TM & MOVE DATA"]
    out = []
    for r in range(2, ws.max_row + 1):
        n = num(cell(ws, r, 1))
        mv = cell(ws, r, 2)
        if n is None and mv is None:
            continue
        out.append(
            {
                "tm": n,
                "move": mv,
                "single": cell(ws, r, 3),
                "infinite": cell(ws, r, 4),
                "category": cell(ws, r, 6),
            }
        )
    return out


def parse_misc(wb):
    ws = wb["Misc."]
    lines = []
    for r in range(1, ws.max_row + 1):
        v = cell(ws, r, 1)
        if v:
            lines.append(str(v))
    return lines


def parse_docx_encounters():
    if not DOCX.exists():
        return []
    try:
        from docx import Document
    except ImportError:
        return []
    doc = Document(DOCX)
    out = []
    section = None
    for para in doc.paragraphs:
        text = para.text.strip()
        if not text:
            continue
        out.append(text)
    return out


def fetch_canonical_ids():
    """Fetch all species name → canonical national-dex id from PokeAPI.
    Adds entries for the romhack's typo'd names. Saved as data/canonical_ids.json
    so the site can map names to PokeAPI sprites/species correctly even when the
    romhack has reshuffled local dex IDs."""
    import urllib.request

    print("Fetching canonical species list from PokeAPI…")
    url = "https://pokeapi.co/api/v2/pokemon-species?limit=1300"
    req = urllib.request.Request(url, headers={"User-Agent": "platinum-redux-docs build"})
    with urllib.request.urlopen(req, timeout=30) as r:
        data = json.load(r)

    out = {}

    def norm(s):
        return re.sub(r"[^a-z0-9]", "", str(s).lower())

    for entry in data["results"]:
        name = entry["name"]  # already lowercase, hyphenated
        m = re.match(r".*/(\d+)/?$", entry["url"])
        if not m:
            continue
        cid = int(m.group(1))
        out[norm(name)] = cid

    # Aliases for romhack typos / hyphenless variants seen in the data
    aliases = {
        "nidoranf": 29, "nidoranm": 32,
        "honchcrow": 430,    # typo: Honchkrow
        "sandlash": 28,      # typo: Sandslash
        "girationa": 487,    # Giratina (Origin form)
        "slowping": 79,      # typo: Slowpoke
        "annihilate": None,  # unknown / not a species
        "corvisquire": 822,  # gen 8 species, valid
        "molders": None,     # unknown
        "pallosand": 770,    # Palossand
        "hooh": 250,
    }
    for k, v in aliases.items():
        if v is not None:
            out[norm(k)] = v
        elif norm(k) in out:
            del out[norm(k)]

    return out


def main():
    wb = load_workbook()
    try:
        canonical = fetch_canonical_ids()
        path = OUT / "canonical_ids.json"
        with open(path, "w", encoding="utf-8") as f:
            json.dump(canonical, f, ensure_ascii=False, separators=(",", ":"))
        print(f"  canonical_ids.json: {path.stat().st_size:,} bytes, {len(canonical)} species")
    except Exception as e:
        print(f"  WARN: could not fetch canonical IDs: {e}")
        print(f"  Site will fall back to romhack local IDs (sprites may be wrong for renumbered species).")

    datasets = {
        "pokemon": parse_pokemon(wb),
        "forms": parse_forms(wb),
        "levelup": parse_levelup(wb),
        "tm_learn": parse_tm_learn(wb),
        "moves": parse_moves(wb),
        "type_chart": parse_type_chart(wb),
        "type_changes": parse_tc_changes(wb),
        "trainers": parse_trainers(wb),
        "bosses": parse_bosses(wb),
        "encounters": parse_encounters(wb),
        "battle_items": parse_battle_items(wb),
        "walkthrough": parse_walkthrough(wb),
        "tm_data": parse_tm_data(wb),
        "credits": parse_misc(wb),
        "encounter_notes": parse_docx_encounters(),
    }

    for key, val in datasets.items():
        path = OUT / f"{key}.json"
        with open(path, "w", encoding="utf-8") as f:
            json.dump(val, f, ensure_ascii=False, indent=2)
        size = path.stat().st_size
        count = len(val) if isinstance(val, list) else (len(val.get("chart", val)) if isinstance(val, dict) else 0)
        print(f"  {key}.json: {size:,} bytes, {count} entries")


if __name__ == "__main__":
    main()
